#!/home/dell/selinum/env/bin/python
"""
Read complete database and convert to .xlsx

Inputs:
    - mysql test databse
oOutput:
    - PartSearchResult_buildteam.xlsx
"""
from openpyxl.compat import range
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
import datetime
import time
import sys
import os
import subprocess

wbq = Workbook()
file_name =  os.path.basename(sys.argv[0]).replace(".py", "")
now = datetime.datetime.now()
aa= now.strftime("%d %H_%M")
datat= datetime.datetime.strptime(aa, '%d %H_%M').strftime('%d %I_%M %p')
file_type = now.strftime('%B')[:3]+ "_" + datat
dest_filenameq = file_name+"_"+file_type +'.xlsx'
ws1q = wbq.active
ws1q.title = "Ipmi test"
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="000000")
border = Border(top=double, left=thin, right=thin, bottom=double)
fill = PatternFill("solid", fgColor="DDDDDD")
fill = GradientFill(stop=("008000", "EEF2EA"))
fillr = GradientFill(stop=("FF0000", "663300"))
font = Font(b=True, color="FFFFFF")
al = Alignment(horizontal="center", vertical="center")
#fill2 = GradientFill(stop=("191970", "0000FF"))  # sdr, fru, event, sel sub heading
fill3 = GradientFill(stop=("FF6633", "FF9900"))  # main 1st heading
fill2 = GradientFill(stop=("FFFF00", "CC9933"))  # main 1st heading

start_time1 = time.time()
sCmdName1 =ws1q
sCmdIPMI1 = 'xlsx_coverage.txt'
numbers = [sCmdIPMI1]
string = [sCmdName1]

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    left1 = Border(left=border.left)
    nu =cell_range.split(":")[1][1:]
    st = cell_range.split(":")[0][:1]
    try:
        en = cell_range.split(":")[1][:1]
    except:
        en = cell_range.split(":")[1][0:2]
    #    print en
    first_cell1 = ws[cell_range.split(":")[0]]
    try:
        for one in range(ord(st), ord(en) +1):
            sheet_n = chr(one) + str(nu)
            first_cell1 = ws[sheet_n]
            if font:
                first_cell1.font = font
    except:
        if font:
            first_cell1.font = font
    if alignment:
        ws.merge_cells(cell_range)
        first_cell1.alignment = alignment
    rows1 = ws[cell_range]
    for cell in rows1[0]:
        cell.border = cell.border + top
    for cell in rows1[-1]:
        cell.border = cell.border + bottom
    for row1 in rows1:
        l1 = row1[0]
        r1 = row1[-1]
        l1.border = l1.border + left
        r1.border = r1.border + right
        if fill:
            for c in row1:
                c.fill = fill

def xlsheet_con():
    global dest_filenameq
    platform_name ="PartSearchResult_buildteam"
    ws1q.title = platform_name
    dest_filenameq = platform_name + '.xlsx'
    for number, names in zip(numbers, string):
        iq = 1
        columsize = 1

        data_s = ['MTrack-ID','Serial Number','Actual location','Part','Revision','User_Defined_Description','Status','Manufacturer','Manufacturer_Part_Number','PartLocation','Frozen_Cost','AgilePartType','CountryofOrigin','Category','TestStatus','PurchasingCC','User_Defined_Original_Cost','AllocatedBorrower',
'Asset','RecallDate','ShippingInfo','BondingInfo','PONumber','Bulk_Qunatity','PartComment','CurrentOwner','Borrower_CC','Borrower','TimesOut','Reason','Due_Date','Last_Inventoried_Date','File_Attachment_Title','System_Service_Tag']
        names.append(data_s)
        style_range(names, 'A' + str(1) + ':AH' + str(1), border=border, fill=fill3, font=font)
       # print "Cmd_test file : -----------------", number
        try:
            with open(number) as fp:
                lines = fp.read().split("\n")
        except:
            print "****** '" + number + "' file is not prasent  "+ number + "in same folde.. **************"
            time.sleep(3)
            sys.exit()
        for cmm_list in range(len(lines)-1):
            data_sp= lines[cmm_list].replace("\n", "").replace('"', '').replace(" ", "").split("#@")
            print data_sp[0], data_sp[1], data_sp[2], data_sp[3], data_sp[4],data_sp[5], data_sp[6],data_sp[7], data_sp[8]
            
            names['A' + str(columsize + 1)] = str(data_sp[0])
            names['B' + str(columsize + 1)] = str(data_sp[1])
            names['C' + str(columsize + 1)] = str(data_sp[2])
            names['D' + str(columsize + 1)] = str(data_sp[3])
            names['E' + str(columsize + 1)] = str(data_sp[4])
            names['F' + str(columsize + 1)] = str(data_sp[5])
            names['G' + str(columsize + 1)] = str(data_sp[6])
            names['H' + str(columsize + 1)] = str(data_sp[7])
            names['I' + str(columsize + 1)] = str(data_sp[8])
            names['J' + str(columsize + 1)] = str(data_sp[9])
            names['E' + str(columsize + 1)] = str(data_sp[10])
            names['K' + str(columsize + 1)] = str(data_sp[11])
            names['L' + str(columsize + 1)] = str(data_sp[12])
            names['M' + str(columsize + 1)] = str(data_sp[13])
            names['N' + str(columsize + 1)] = str(data_sp[14])
            names['O' + str(columsize + 1)] = str(data_sp[15])
            names['P' + str(columsize + 1)] = str(data_sp[16])
            names['Q' + str(columsize + 1)] = str(data_sp[17])
            names['R' + str(columsize + 1)] = str(data_sp[18])
            names['S' + str(columsize + 1)] = str(data_sp[19])
            names['T' + str(columsize + 1)] = str(data_sp[20])
            names['U' + str(columsize + 1)] = str(data_sp[21])
            names['V' + str(columsize + 1)] = str(data_sp[22])
            names['W' + str(columsize + 1)] = str(data_sp[23])
            names['X' + str(columsize + 1)] = str(data_sp[24])
            names['Y' + str(columsize + 1)] = str(data_sp[25])
            names['Z' + str(columsize + 1)] = str(data_sp[26])

            names['AA' + str(columsize + 1)] = str(data_sp[27])
            names['AB' + str(columsize + 1)] = str(data_sp[28])
            names['AC' + str(columsize + 1)] = str(data_sp[29])
            names['AD' + str(columsize + 1)] = str(data_sp[30])
            names['AE' + str(columsize + 1)] = str(data_sp[31])
            names['AF' + str(columsize + 1)] = str(data_sp[32])
            names['AG' + str(columsize + 1)] = str(data_sp[33])
            names['AH' + str(columsize + 1)] = str(data_sp[34])

            colour = columsize
            iq +=1
            columsize += 1
            style_range(names, 'A' + str(colour + 1) + ':AH' + str(colour + 1), border=border, fill=fill2, font=font)
            wbq.save(filename=dest_filenameq)

    wbq.save(filename=dest_filenameq)
    time.sleep(2)
    subprocess3 = subprocess.check_output('cp -rf  /home/narayana/workspace/M-Track-Pipeline/PartSearchResult_buildteam.xlsx  /opt/lampp/htdocs/', shell=True)
    print subprocess3

def main():
    xlsheet_con()

if __name__ == '__main__':
    main()
